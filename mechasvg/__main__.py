import os, random, datetime, functools, math, sys
try:
	import tkinter as tk
	from tkinter import ttk, filedialog, messagebox
except ImportError as message:
	print(message)
	print()
	print("Please make sure you have python3, tkinter and ttk installed")
	input("Press enter to leave")
	exit()


class Preferences:
	def __init__(self):
		################################ RELATIVELY STRAIGHTFORWARD MODIFICATIONS ######################################
		# number of structures for each path
		self.n_structures = 19 #max = 25 for random catalytic cycle generator
		#run command
		self.n_connectors = 8
		self.windons_command = "start inkscape.exe ./.E_profile.svg" # Please note .E_profile.svg is a hidden file!
		self.linux_command = "inkscape ./.E_profile.svg" # Please note .E_profile.svg is a hidden file!
		self.command_line = self.windons_command if os.name == "nt" else self.linux_command
		# SVG colors
		self.menu_a = ["grey","black","blue","darkblue","red","darkred","green","darkgreen"]
		#SVG line widths
		self.menu_b = ["2","3","4","5","6"]
		#SVG line stiles
		self.svg_repl = {"full": "","dashed":'stroke-dasharray="10,10"',"dashed1":'stroke-dasharray="6,6"'}
		# Random catalytic cycle generator
		self.trickster = True # Include random catalytic cycle generator?
		self.name = "MechaSVG v 0.0.5"
		######################## YOU ARE PROBABLY BETTER OFF NOT MESSING WITH THE FOLLOWING ############################
		self.menu_c = list(self.svg_repl.keys())
		# TDI and TDTS placement corrections

		self.placement = {
			"Top":[[-37,-22,-7],[-22,-7,0]],
			 "Middle"  :[[-5,15,30],[-5,15,30]],
			 "Bottom" :[[17,32,47],[17,32,0]]
			}
		self.alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
		self.menu_d = list(self.placement.keys())
		self.menu_e = [self.alphabet[n] for n in range(8)] #Will change the number of Paths available
		self.menu_f = ["opt_{}".format(a.lower()) for a in self.menu_e]
		self.menu_g = ["tab_{}".format(a.lower()) for a in self.menu_e]
		self.menu_h = ["Path {}".format(a) for a in self.menu_e]
		self.menu_i = ["#{}".format(a) for a in self.menu_e]
		self.menu_z = [" ","TS","INT"]
		try:
			import openpyxl
			self.xlsx = True
		except ImportError:
			self.xlsx = False
		if self.xlsx:
			self.allowed_extensions = {
				"title": "Save state",
				"defaultextension": ".xlsx",
				"filetypes": [("Spreadsheet", ".xlsx"), ("Saved State File", ".ssf"),  ("Text file", ".txt")]}
		else:
			self.allowed_extensions = {
				"title": "Save state",
				"defaultextension": ".ssf",
				"filetypes": [("Saved State File", ".ssf"), ("Text file", ".txt")]}
		filename = sys.argv[-1]
		ext = [".xlsx",".ssf",".txt"] if self.xlsx else [".ssf",".txt"]
		if os.path.isfile(filename) and any(filename.endswith(a) for a in ext):
			self.filename = filename
class State:
	def __init__(self, text):
		self.list = [a.split("/|") for a in text]
		for a,b in zip(pref.menu_f,pref.menu_g): setattr(self,a,[]);setattr(self,b,[])
		self.con = []
		self.message = []
		tab, opt, con = [], [], []
		for i,a in enumerate(self.list):
			if len(a) == 5:	tab.append(a)
			elif len(a) == 3: opt.append(a)
			elif len(a) == 1:
				if a[0] == "#CON":
					if len(con) > pref.n_connectors:
						self.message.append("Exceeding number of conectors")
					self.con = con[:min(len(con),pref.n_connectors)]
					break
				if not a[0] in ["#{}".format(n) for n in pref.menu_e]:
					if a[0] in ["#{}".format(n) for n in pref.alphabet]:
						self.message.append("Exceeding number of paths")
					continue
				for b,c,d in zip(pref.menu_i,pref.menu_g,pref.menu_f):
					if a[0] == b:
						if len(tab) > pref.n_structures:
							self.message.append("Exceding number of structures")
						setattr(self,c,tab[:min(len(tab),pref.n_structures)])
						setattr(self,d,opt)
						tab, opt = [],[]
			elif len(a) == 7: con.append(a)
	def print_interpretation(self):
		for a in vars(self):
			print(a,"---->",getattr(self,a))

class Note(ttk.Notebook):
	def __init__(self,parent,*args,**kwargs):
		ttk.Notebook.__init__(self,parent,*args,**kwargs)
		for a,b in zip(pref.menu_g,pref.menu_h):
			setattr(self, a,TabFramePaths(self,name=b))
		self.tab_connections = TabFrameConnections(self,name="Connections")
		self.grid(column=0,row=0,sticky='news')
		self.grid_columnconfigure(0, weight=1)

class TabFramePaths(ttk.Frame):
	def __init__(self,parent,name,*args,**kwargs):
		ttk.Frame.__init__(self,parent,*args,**kwargs,height=50)
		self.parent = parent
		self.parent.add(self, text=name)
		#########
		self._build_options()
		#######
		for n in range(3):
			self.grid_columnconfigure(n+2, weight=[2,1,1][n])
		#########
		for a,b in zip([1,2,3,4,5,7],["Type",'Structure Name','Free Energy','Entalphy',"Move",'Alignment']):
			label = tk.Label(self, text=b)
			if a == 5: label.grid(column=a, row=9, rowspan=1, columnspan = 2,sticky="news")
			else: label.grid(column=a, row=9, rowspan=1,sticky="news")
		#########
		self.data = [[None,None,None,None,None] for _ in range(pref.n_structures)]
		#########
		for n in range(pref.n_structures):
			label = tk.Label(self, text='#{}'.format(n+1))
			label.grid(column=0, row=10+n, rowspan=1)
			for b in [1,2,3]:
				self.data[n][b] = tk.Entry(self,justify=tk.CENTER,bd=2,width=10 if b in [2,3] else 15)
				self.data[n][b].insert(0,"")
				self.data[n][b].grid(column=1+b, row=10+n,padx="0",sticky="news")
			if not n == 0:
				button = tk.Button(self, text=u'\u2191', command=lambda x = n: self._move(x,x-1), padx="1")
				button.config(width=1)
				button.grid(column=5, row=10 + n)
			if not n+1 == pref.n_structures:
				button = tk.Button(self, text=u'\u2193', command=lambda x = n: self._move(x,x+1), padx="1")
				button.config(width=1)
				button.grid(column=6, row=10 + n)
			self.data[n][0] = tk.StringVar()
			menu = tk.OptionMenu(self,self.data[n][0],*pref.menu_z)
			self.data[n][0].set(pref.menu_z[0])
			menu.config(width="3")
			menu.grid(column=1, row=10 + n)
			self.data[n][4] = tk.StringVar()
			menu = tk.OptionMenu(self,self.data[n][4],*pref.menu_d)
			self.data[n][4].set(pref.menu_d[1])
			menu.config(width="6")
			menu.grid(column=7, row=10 + n)
	def _build_options(self):
		self.option_menu = ttk.Frame(self)
		setattr(self.option_menu,"line_opt_data",[[None,None,None],[None,None,None]])
		for n,name in enumerate(["Main[Color/Width/Strike]","Link[Color/Width/Strike]"]):
			box = self.boxify(self.option_menu,name=name,column=n)
			for a in range(3):box.grid_columnconfigure(a, weight=1)
			for a,b in zip([0,1,2],[pref.menu_a,pref.menu_b,pref.menu_c]):
				self.option_menu.line_opt_data[n][a] = tk.StringVar()
				color_menu = tk.OptionMenu(box, self.option_menu.line_opt_data[n][a],*b)
				self.option_menu.line_opt_data[n][a].set(b[[1,1,0][a] if n == 0 else [1,0,2][a]])
				color_menu.config(width=["9","1","7"][a])
				color_menu.grid(column=a,row=0,sticky="news")
		self.option_menu.grid(column=0,row=0,columnspan = 8,rowspan=8,sticky="news")
		self.option_menu.grid_columnconfigure([0,1], weight=1)
	def _move(self,n,x):
		line_n = [a.get() for a in self.data[n]]
		other = [a.get() for a in self.data[x]]
		for i, a in enumerate(other):
			if type(self.data[n][i]) is tk.Entry:
				self.data[n][i].delete(0,tk.END)
				self.data[n][i].insert(0,a)
			elif type(self.data[n][i]) is tk.StringVar:
				self.data[n][i].set(a)
		for i, a in enumerate(line_n):
			if type(self.data[x][i]) is tk.Entry:
				self.data[x][i].delete(0,tk.END)
				self.data[x][i].insert(0,a)
			elif type(self.data[x][i]) is tk.StringVar:
				self.data[x][i].set(a)
	@staticmethod
	def boxify(other, name, column):
		box = ttk.LabelFrame(other, text=name)
		box.grid(column=column, row=0, sticky="news")
		return box
class TabFrameConnections(ttk.Frame):
	def __init__(self,parent,name,*args,**kwargs):
		ttk.Frame.__init__(self,parent,*args,**kwargs)
		self.parent = parent
		self.parent.add(self, text=name)
		self.data = [[None,None,None,None,None,None,None] for _ in range(pref.n_connectors)]
		for n in range(pref.n_connectors):
			con = tk.LabelFrame(self,text="Conector {}".format(n+1))
			con.grid(column=0, row=n*2,  columnspan=5, pady="0",padx="2", rowspan=2,sticky="news")
			for b in range(2):
				label = tk.Label(con, text="From path" if b ==0 else "to path")
				label.grid(column=b*4+0, row=0, sticky="w")
				self.data[n][b*2] = tk.StringVar()
				menu = tk.OptionMenu(con,self.data[n][b*2],*pref.menu_e)
				menu.config(width="2")
				menu.grid(column=b*4+1, row=0,sticky = "e")
				label = tk.Label(con, text=", number" if b ==0 else ", number")
				label.grid(column=b * 4 + 2, row=0, sticky="w")
				self.data[n][b*2+1] = tk.StringVar()
				menu = tk.OptionMenu(con,self.data[n][b*2+1],*[a+1 for a in range(pref.n_structures)])
				menu.config(width="2")
				menu.grid(column=b*4+3, row=0,sticky = "e")
			label = tk.Label(con,text="Color/Width/Strike:")
			label.grid(column=0,row=1,columnspan=2)
			for a, b in zip([0, 1, 2], [pref.menu_a, pref.menu_b, pref.menu_c]):
				self.data[n][a+4] = tk.StringVar()
				self.color_menu = tk.OptionMenu(con, self.data[n][a+4], *b)
				self.data[n][a+4].set(b[0])
				self.color_menu.config(width="12")
				self.color_menu.grid(column=2 + a*2,columnspan =2, row=1)
class GeneralMenu(tk.LabelFrame):
	def __init__(self,parent,name,*args,**kwargs):
		ttk.LabelFrame.__init__(self,parent,text=name,*args,**kwargs)
		self.include = []
		self.main = []
		self.span = []
		self.titles = ["Main Title", "ΔE in kcal·mol⁻¹", "Reaction coordinate"]
		self.command = ""
		self.aesthetics = []
		###BUILD
		self._build_all()
		if hasattr(pref,"filename"):
			self.f = pref.filename
			self.load_state(getattr(pref,"filename"))
	def _build_all(self):
		self.note = ttk.Notebook(self.boxify("Advanced options", 2))
		self.note.grid(column=0, row=0, sticky="news")
		self.note.grid_columnconfigure(0, weight=1)
		self._change_win_title("Unsaved")
		self._build_other_opt()
		self._build_span_opt()
		self._build_aesthetics()
		self._build_path_sel()
		self._build_titles(6)
		self._build_loadsave(7)
		if pref.trickster: self._build_generator(8)
		self._build_preview(9)
		self._build_message(10)
	def _build_path_sel(self):
		box = self.framefy("Include")
		for i,a in enumerate([*pref.menu_h,"Connections"]):
			self.include.append(tk.IntVar(value=1))
			c1 = tk.Checkbutton(box, text=a, variable=self.include[i], onvalue=1, offvalue=0)
			c1.grid(column=i%6,row=i//6)
	def _build_other_opt(self):
		box = self.framefy("Main")
		options = ["Use enthalpy instead of free energy","Use comma as decimal","Include complementary (H or G values)"]
		for i,a in enumerate(options):
			self.main.append(tk.IntVar(value=0))
			c1 = tk.Checkbutton(box, text=a, variable=self.main[i], onvalue=1, offvalue=0)
			c1.grid(column=i%2,row=i//2,sticky="w")
	def _build_span_opt(self):
		box = self.framefy("Span")
		box.grid_columnconfigure(1, weight=1)
		options = ["Atempt span","Irrespective of type (TS/INT)","Big arrow"]
		for i,a in enumerate(options):
			self.span.append(tk.IntVar(value=0))
			c1 = tk.Checkbutton(box, text=a, variable=self.span[i], onvalue=1, offvalue=0)
			c1.grid(column=i*2,row=0,columnspan=2,sticky="news")
		label = tk.Label(box, text="Input units:")
		label.grid(column=0, row=1,sticky="news")
		self.span.append(tk.StringVar())
		options = ["kcal/mol","kJ/mol"]
		menu = tk.OptionMenu(box, self.span[-1], *options)
		self.span[-1].set(options[0])
		menu.config(width="8")
		menu.grid(column=1, row=1,sticky="w")
		label = tk.Label(box, text="Temperature (°C):")
		label.grid(column=2, row=1, columnspan = 2)
		self.span.append(tk.Entry(box, justify=tk.CENTER, bd=3, width=6))
		self.span[-1].insert(0, "25")
		self.span[-1].grid(column=4, row=1, padx="3",pady="4", sticky="news")
	def _build_aesthetics(self):
		box = self.framefy("Aesthetics")
		a = ["   ", "( )", "[ ]", r"{ }", '" "', "' '"]
		b = [a * 2 for a in range(11)]
		c = [1,2]
		e = [a,a,b,c]
		f = ["G:","H:","Width:","Decimal"]
		for a,(b,c) in enumerate(zip(f,e)):
			label = tk.Label(box,text=b)
			label.grid(column=2*a,row=0)
			self.aesthetics.append(tk.StringVar())
			menu = tk.OptionMenu(box, self.aesthetics[a], *c)
			menu.config(width="2")
			menu.grid(column=a * 2 + 1, row=0)
		for i,a in enumerate(["   ","( )",10,1]):
			self.aesthetics[i].set(a)
		a = [0 + a * 10 for a in range(11)]
		b = [60 + a * 5 for a in range(11)]
		c = [" ","‡ (big)","‡ (small)"]
		d = ["X offset:","X dist:","TS mark:"]
		e = [a,b,c]
		for a, (b, c) in enumerate(zip(d, e)):
			label = tk.Label(box, text=b)
			label.grid(column=0 if a ==0 else 2 * a +1, row=1, columnspan = 2 if a ==0 else 1)
			self.aesthetics.append(tk.StringVar())
			menu = tk.OptionMenu(box, self.aesthetics[a+4], *c)
			menu.config(width="8" if a == 2 else "2")
			menu.grid(column=a * 2 + 2, row=1, columnspan = 3 if a == 2 else 1, sticky="news" if a==2 else "")
		for i, a in enumerate([40, 80, " "]):
			self.aesthetics[i+4].set(a)
	def _build_titles(self,idx):
		box = self.boxify("Titles",idx)
		for a,b,c in zip([0,1,2],self.titles,["Main:","y:","x:"]):
			label = tk.Label(box, text=c,width=10)
			label.grid(column=0, row=a)
			self.titles[a] = tk.Entry(box, justify=tk.CENTER, bd=2, width=50)
			self.titles[a].insert(0, b)
			self.titles[a].grid(column=1, row=a, padx="0", sticky="news")
	def _build_loadsave(self,idx):
		box = self.boxify("Close, Load & Save States", idx)
		label = tk.Label(box, text="Path's and connection's info")
		label.grid(column=0,row=0,sticky="w")
		button = tk.Button(box, text="Close", command=self._blank_state, padx="1")
		button.config(width=7)
		button.grid(column=1, row=0, sticky="e")
		button = tk.Button(box, text="Load", command=self.load_state, padx="1")
		button.config(width=7)
		button.grid(column=2,row=0,sticky="e")
		button = tk.Button(box, text="Save as", command=self._save_as, padx="1")
		button.config(width=7)
		button.grid(column=3,row=0,sticky="e")
		button = tk.Button(box, text="Save", command=self._save, padx="1")
		button.config(width=7)
		button.grid(column=4,row=0,sticky="e")
		box.grid_columnconfigure(0, weight=1)
	def _build_generator(self,idx):
		box = self.boxify("Generate random catalytic cycle", idx)
		label = tk.Label(box, text="Random catalytic cycle generator")
		label.pack(side=tk.LEFT)
		button = tk.Button(box, text="Fill in data", command=self._ask_confirmation, padx="1")
		button.config(width=10)
		button.pack(side=tk.RIGHT)
	def _build_preview(self,idx):
		box = self.boxify("Preview with either command or default manager, or save svg file", idx)
		self.command = tk.Entry(box, justify=tk.CENTER, bd=4)
		self.command.insert(0, pref.command_line)
		self.command.grid(column=0,row=0,sticky="news")
		button = tk.Button(box, text="Command", command=self.run_data_a, padx="1")
		button.config(width=8)
		button.grid(column=1,row=0,sticky="e")
		button = tk.Button(box, text="Default", command=self.run_data_b, padx="1")
		button.config(width=8)
		button.grid(column=2,row=0,sticky="e")
		button = tk.Button(box, text="Save svg", command=self.return_svg, padx="1")
		button.config(width=8)
		button.grid(column=3,row=0,sticky="e")
		box.grid_columnconfigure(0, weight=1)
	def _build_message(self,idx):
		box = self.boxify("Message",idx)
		m = tk.Message(box)
		scrollbar = tk.Scrollbar(box)
		scrollbar.grid(column=1,row=0,sticky="nes")
		self.msg = tk.Text(box,width=40,yscrollcommand=scrollbar.set,height=12 if pref.trickster else 18,state="disabled",background=m.cget("background"),relief="flat",wrap=tk.WORD,font=("Helvetica", 8))
		m.destroy()
		self.msg.grid(column=0,row=0,sticky="news")
		if not pref.xlsx:
			self.message("Welcome!\n\nTo enable .xlsx file support, please install openpyxl python library via the shell command:\npython3 -m pip install openpyxl")
		else:
			self.message("Welcome!")
		box.grid_rowconfigure(0, weight=1)
		box.grid_columnconfigure(0, weight=1)
		self.grid_rowconfigure(idx, weight=1)
	def _save(self,ignore=False):
		#TODO
		try:
			if self.f.endswith(".xlsx") and pref.xlsx:
				from openpyxl import Workbook
				wb = Workbook()
				wb.remove(wb.active)
				for a,b in zip(self.gen_data(type=".xlsx"),pref.menu_e):
					sheet = wb.create_sheet(title=f'Path {b}')
					for i,c in enumerate(a,start=1):
						sheet.append(c[1:4])
				try:
					wb.save(self.f)
				except PermissionError:
					self.message(f"Error while saving file!\nIs the file:\n'{self.f}' already open?")
			else:
				try:
					with open(self.f,"w") as out:
						if self.f.endswith(".ssf"):
							txt = "".join(a + "/$" + "\n" for a in self.gen_data())
							out.write(txt)
						elif self.f.endswith(".txt"):
							txt = "\n".join(a for a in self.gen_data(type=".txt") if len(a.split()) >= 1)
							out.write(txt)
				except PermissionError:
					self.message(f"Error while saving file!\nIs the file:\n'{self.f}' already open?")
		except AttributeError: self._save_as()
		except FileNotFoundError: self._save_as()
	def _save_as(self):
		self.f = tk.filedialog.asksaveasfilename(**pref.allowed_extensions)
		self._change_win_title(self.f)
		if any(self.f.endswith(a) for a in (".sff",".txt",".xlsx")):self._save()

	def load_state(self,file_n=None):
		if file_n is None:
			file_n = tk.filedialog.askopenfilename(**pref.allowed_extensions)
		try:
			if file_n.endswith(".xlsx") and pref.xlsx:
				self._blank_state(ask=False)
				import openpyxl
				try:
					wb = openpyxl.load_workbook(file_n)
				except:
					self.message(f"Could not read {file_n} as xlsx file!\nAre you sure this is a proper xlsx file?")
					return
				notes = [getattr(note, a) for a in pref.menu_g]
				exceeded = False
				for a,b in zip(wb.sheetnames, notes):
					sheet = wb[a]
					for n in range(1,pref.n_structures+10):
						if n > pref.n_structures:
							if any(sheet.cell(row=n,column=i).value is None for i in range(1,4)):
								exceeded = True
							continue
						for i in range(1,4):
							if sheet.cell(row=n,column=i).value is None: continue
							b.data[n-1][i].insert(0,str(sheet.cell(row=n,column=i).value))
				if exceeded:
					self.message("Exceeding number of structures")
			else:
				with open(file_n, mode="r") as file:
					if file_n.endswith(".ssf"):
						self._blank_state(ask=False)
						state = "".join(file.read().splitlines()).split("/$")
						try:
							state = State(state)
							notes = [getattr(note, a) for a in pref.menu_g]
							dat_states = [getattr(state, a) for a in pref.menu_g]
							opt_states = [getattr(state, a) for a in pref.menu_f]
							for a, b, c in zip(notes, dat_states, opt_states):
								for n in range(3):
									for d in range(2):
										a.option_menu.line_opt_data[d][n].set(c[d][n])
								for i, line in enumerate(b):
									for n in range(3):
										a.data[i][n + 1].delete(0, tk.END)
										a.data[i][n + 1].insert(0, line[n + 1])
									a.data[i][0].set(line[0])
									a.data[i][4].set(line[4])
							for i, line in enumerate(state.con):
								for n in range(7):
									note.tab_connections.data[i][n].set(line[n])
							if state.message:
								self.message("\n".join(state.message))
						except IndexError:
							pass

					elif file_n.endswith(".txt"):
						self._blank_state(ask=False)
						all_tabs = {}
						tab_data = []
						for line in file.read().splitlines():
							line = line.split()
							non_hash = True if len(line) == 1 and not line[0].startswith("#") else False
							if len(line) >= 2 or non_hash:
								tab_data.append(line)
							elif len(line) == 1 and any(line[0] == f"#{a}" for a in pref.menu_e):
								all_tabs[line[0]] = tab_data
								tab_data = []
						if len(all_tabs) == 0 and len(tab_data) != 0:
							all_tabs["#A"] = tab_data
						missing = [b for b in [f"#{a}" for a in pref.menu_e] if b not in all_tabs.keys()]
						for a in missing: all_tabs[a] = []
						notes = [getattr(note, a) for a in pref.menu_g]
						exceeded = False
						for a,b in zip(notes,sorted(all_tabs.keys())):
							for i,c in enumerate(all_tabs[b]):
								if i >= pref.n_structures:
									exceeded = True
									continue
								for n in range(3):
									try: a.data[i][n+1].insert(0,c[n])
									except IndexError: pass
						if exceeded:
							self.message("Exceeding number of structures")
					else:
						return
		except FileNotFoundError:
			self.message("File not found!")
			return
		finally:
			self._change_win_title(file_n)
			self.f = file_n

	def fill_in(self):
		size = random.random()
		max_value = min(len(pref.alphabet), pref.n_structures)
		lenght = random.randint(1,max_value)
		tab = getattr(note,[a for a in pref.menu_g][note.index(note.select())])
		for i,n in zip(range(max_value),pref.alphabet):
			value = size*random.randrange(-100,100)
			for idx in range(len(tab.data[i])):
				if idx == 1:
					tab.data[i][idx].delete(0, tk.END)
					if i+1 < lenght: tab.data[i][idx].insert(0, n)
					elif i + 1 == lenght: tab.data[i][idx].insert(0, "A'")
				elif idx == 2:
					tab.data[i][idx].delete(0, tk.END)
					if i < lenght: tab.data[i][idx].insert(0,"{:.2f}".format(value))
				elif idx == 3:
					tab.data[i][idx].delete(0, tk.END)
					if i < lenght: tab.data[i][idx].insert(0,"{:.2f}".format(value + random.choice([-random.random(), +random.random()])))
				elif idx == 4:
					tab.data[i][idx].set(pref.menu_d[1])
		max_v, min_v = None, None
		for i in range(max_value):
			if max_v is None: max_v = [i,tab.data[i][2].get()]
			if min_v is None: min_v = [i,tab.data[i][2].get()]
			if i < lenght and float(tab.data[i][2].get()) > float(max_v[1]): max_v = [i, tab.data[i][2].get()]
			if i < lenght and float(tab.data[i][2].get()) < float(min_v[1]) : min_v = [i, tab.data[i][2].get()]
			if i == 0: tab.data[i][0].set("INT")
			elif i+1 == lenght: tab.data[i][0].set("INT")
			elif i >= lenght: tab.data[i][0].set("  ")
			else:
				if float(tab.data[i-1][2].get()) < float(tab.data[i][2].get()) > float(tab.data[i+1][2].get()):
					tab.data[i][0].set("TS")
				else:
					tab.data[i][0].set("INT")
		tab.data[max_v[0]][4].set(pref.menu_d[2])
		tab.data[min_v[0]][4].set(pref.menu_d[0])
	def _ask_confirmation(self):
		if note.index(tk.END) + -1 <= note.index(note.select()):
			self.message("Cannot fill in data for connection tab!\n")
			return
		msgbox = tk.messagebox.askquestion(
			f'Fill in random catalytic cycle at {pref.menu_h[note.index(note.select())]}',
			'Are you sure? All unsaved data will be lost!', icon='warning')
		if msgbox == "yes":
			self.fill_in()
			self._change_win_title("Unsaved")
			if hasattr(self,"f"): del(self.f)
	def _change_win_title(self,path):
		window.title(f"{pref.name} @ {path}")
	def _blank_state(self,ask=True):
		if ask:
			msgbox = tk.messagebox.askquestion('Close document', 'Are you sure? All unsaved data will be lost!', icon='warning')
			if msgbox != "yes":return
			self._change_win_title("Unsaved")
			if hasattr(self,"f"): del(self.f)
		for a in [getattr(note,a) for a in pref.menu_g]:
			for i in range(pref.n_structures):
				for idx in range(5):
					if idx == 0: a.data[i][idx].set(pref.menu_z[0])
					if idx in [1,2,3]: a.data[i][idx].delete(0, tk.END)
					if idx  == 4: a.data[i][idx].set(pref.menu_d[1])
		for a in [getattr(note,a) for a in pref.menu_g]:
			for n in range(2):
				for idx, b in zip([0, 1, 2], [pref.menu_a, pref.menu_b, pref.menu_c]):
					a.option_menu.line_opt_data[n][idx].set(b[[1,1,0][idx] if n == 0 else [1,0,2][idx]])
		for a in range(pref.n_connectors):
			for b in range(4):
				note.tab_connections.data[a][b].set("")
			for b,c in zip(range(3),[pref.menu_a, pref.menu_b, pref.menu_c]):
				note.tab_connections.data[a][b+4].set(c[0])
	def message(self,text):
		now = datetime.datetime.now()
		self.msg.configure(state="normal")
		self.msg.tag_add("start", "0.0", tk.END)
		self.msg.tag_config("start", foreground="grey")
		if type(text) == str: text = [text]
		for txt in text:
			self.msg.insert("1.0",txt+"\n")
		self.msg.insert("1.0", "[" + ":".join(["{:02d}".format(a) for a in [now.hour, now.minute, now.second]]) + "] "+"\n")
		self.msg.configure(state="disabled")
	def boxify(self,name,row):
		box = ttk.LabelFrame(self, text=name)
		box.grid(column=0, row=row, sticky="news")
		box.grid_columnconfigure(0, weight=1)
		return box
	def framefy(self,name):
		x = tk.Frame()
		x.grid(column=0, row=0, sticky="news")
		x.grid_columnconfigure(0, weight=1)
		self.note.add(x,text=name)
		return x
	def print_data(self):
		notes = [getattr(note,a) for a in pref.menu_g]
		for a,b in zip(notes,pref.menu_e):
			print(f"NOTE {b}")
			for idx,line in enumerate(getattr(a,"data")):
				if any(c.get().strip() != "" for c in line[:-1]):
					print(f"#{idx+1}",[n.get() for n in line])
		print("NOTE CONNECTIONS")
		for idx,a in enumerate(note.tab_connections.data):
			if any(c.get().strip() != "" for c in a[:-3]):
				print(f"#{idx+1}", [n.get() for n in a])
	def gen_data(self,type=".ssf"):
		notes = [getattr(note,a) for a in pref.menu_g]
		ssf_data = []
		txt_data = []
		xlsx_data = []
		for a,b in zip(notes,pref.menu_e):
			xlsx = []
			for idx,line in enumerate(getattr(a,"data")):
				c = [n.get() for n in line]
				assert all("/|" not in d for d in c), self.message("'/|' is not allowed in names or energy values")
				assert all("/$" not in d for d in c), self.message("'/$' is not allowed in names or energy values")
				ssf_data.append("/|".join(c))
				txt_data.append("{:<20} {:>10} {:>10}".format(*c[1:4]))
				xlsx.append(c)
			ssf_data.append("/|".join([d.get() for d in a.option_menu.line_opt_data[0]]))
			ssf_data.append("/|".join([d.get() for d in a.option_menu.line_opt_data[1]]))
			ssf_data.append("#{}".format(b))
			txt_data.append("#{}".format(b))
			xlsx_data.append(xlsx)
		for idx,a in enumerate(note.tab_connections.data):
			ssf_data.append("/|".join([n.get() for n in a]))
		ssf_data.append("#CON")
		if type == ".ssf":
			return ssf_data
		elif type == ".txt":
			return txt_data
		elif type == ".xlsx":
			return xlsx_data

	def save_svg_as(self):
		return tk.filedialog.asksaveasfilename(defaultextension=".svg", title="Save svg", filetypes=[("Scalable Vector Graphics", ".svg")])
	def run_data_a(self):
		self.return_svg(promp=False); os.system(self.command.get())
	def run_data_b(self):
		self.return_svg(promp=False); os.startfile(os.path.join(os.getcwd(), ".E_profile.svg"))
	def return_svg(self,promp=True):
		svg_name = None if promp == False else self.save_svg_as()
		kwargs = {
			"state": State(self.gen_data()), #ok
			"include" : [a.get() for a in self.include], #ok
			"title" : {a:b.get() for a,b in zip(["main","y","x"],self.titles)},#ok
			"main" : {a: b.get() for a, b in zip(["energy", "comma", "include"], self.main)}, #ok
			"aesthetics" : {a:b.get() for a,b in zip(["g","h","width","decimal","offset","distance","mark"],self.aesthetics)},#ok
			"span" : {a:b.get() for a,b in zip(["span","irrespective","big_arrow","units","temperature"],self.span)}
		}
		msg = SvgGenEsp(**kwargs).save_svg(svg_name)
		if not msg is None: self.message(msg)

class SvgGenEsp:
	def __init__(self,state,include,title,main,aesthetics,span):
		self.span_worthy = True
		self.msg = []
		self.aesthetics = aesthetics
		self.wide = [20-int(aesthetics["width"]),40+int(aesthetics["width"])]
		self.include = include
		self.span = span
		self.e_source = 4 if main["energy"] == 1 else 3
		self.e_complement = 3 if self.e_source == 4 else 4
		self.title = title
		self.span_request = True if span["span"] == 1 else False
		self.comma = True if main["comma"] == 1 else False
		self.include_np = True if main["include"] == 1 else False
		if self.span_request:
			self.temperature = self._verify_temp(span["temperature"])
		self.big_arrow = True if span["big_arrow"] == 1 else False
		################################################################################################################
		self.state = state
		self.tab_v = ["tab_{}_v".format(a.lower()) for a in pref.menu_e]
		for a,b in zip(self.tab_v,pref.menu_g):
			d = [[i + 1, *c] for i, c in enumerate(getattr(self.state,b)) if is_str_float(c[self.e_source-1].replace(",","."))]
			d = [[float(c.replace(",", ".")) if i == self.e_source else c for i, c in enumerate(e)] for e in d]
			setattr(self,a,d)
		self.data_atr = [a for a, b in zip(self.tab_v, self.include) if b == 1 and getattr(self,a)]
		self.opt_atr = [{a:b for a,b in zip(self.tab_v,pref.menu_f)}[a] for a in self.data_atr]
		self.paths = []
		self.svg_code = ['<?xml version="1.0" encoding="UTF-8" ?>']


	def _verify_temp(self,value):
		if is_str_float(value.replace(",",".")):
			if float(value) <= -273.15:
				self.span_worthy = False
				self.msg.append("Temperature should not be smaller than or equal to -273{}15 °C\n".format("," if self.comma else "."))
				return None
			else:
				t = float(value) + 273.15
				self.msg.append("Temperature is set to {} K\n".format(self.commafy("{:.2f}".format(t))))
				return t
		else:
			self.span_worthy = False
			self.msg.append("Unrecognized temperature: {}\n".format("value"))
	def commafy(self,item):
		return str(item).replace(".", ",") if self.comma else str(item).replace(",", ".")
	@functools.lru_cache(maxsize=1)
	def max_value(self):
		return max(max(a[self.e_source] for a in getattr(self,atr)) for atr in self.data_atr)
	@functools.lru_cache(maxsize=1)
	def min_value(self):
		return min(min(a[self.e_source] for a in getattr(self,atr)) for atr in self.data_atr)
	@functools.lru_cache(maxsize=1)
	def delta_value(self):
		return self.max_value()-self.min_value()
	@functools.lru_cache(maxsize=1)
	def n_col(self):
		try: x = max(max(a[0] for a in getattr(self, atr) if a) for atr in self.data_atr)
		except ValueError: x = 0
		return x
	@functools.lru_cache(maxsize=1)
	def set_height(self):
		self.paths = []
		for idx_a, a in enumerate(getattr(self,atr) for atr in self.data_atr):
			path = []
			for idx_b, b in enumerate(a):  # For every structure
				try: height = int(round(abs(400 - (b[self.e_source] - self.min_value()) * 400 / self.delta_value())))
				except ZeroDivisionError: height = 250
				path.append([*b,height])
			self.paths.append(path)
		return self.paths
	@functools.lru_cache(maxsize=1)
	def graph_frame(self):
		a = [
			'<svg width="{0}" viewBox="30 0 {0} 500" height="500" xmlns="http://www.w3.org/2000/svg">',
			'    <line x1="100" y1="25" x2="100" y2="475" stroke="black" stroke-width="2"/>',
			'    <line x1="100" y1="475" x2="{}" y2="475" stroke="black" stroke-width="2"/>',
			'    <text x="{}" y="20" font-size="22" text-anchor="middle" fill="black">{}</text>',
			'    <text x="-250" y="55" font-size="22" {} text-anchor="middle" fill="black">{}</text>',
			'    <text x="{}" y="495" font-size="22" text-anchor="middle" fill="black">{}</text>']
		a[0] = a[0].format(self.n_col() * int(self.aesthetics["distance"]) + int(self.aesthetics["offset"]) + 100)
		a[2] = a[2].format(self.n_col() * int(self.aesthetics["distance"]) + int(self.aesthetics["offset"]) + 75)
		a[3] = a[3].format(int(self.n_col() * 40 + int(self.aesthetics["distance"])), self.title["main"].encode("ascii", "xmlcharrefreplace").decode("utf-8"))
		a[4] = a[4].format('transform="rotate(-90)"', self.title["y"].encode("ascii", "xmlcharrefreplace").decode("utf-8"))
		a[5] = a[5].format(int(self.n_col() * 40 + int(self.aesthetics["distance"])), self.title["x"].encode("ascii", "xmlcharrefreplace").decode("utf-8"))
		self.svg_code.extend(a)

	@functools.lru_cache(maxsize=1)
	def graph_grid(self):
		step_size = max(a if 10 * a < abs(self.delta_value()) else 0.05 for a in [0.1,0.2, 0.5, 1, 2, 5, 10, 25, 100])
		max_e = round((math.ceil(self.max_value() + self.delta_value()) / 10)) * 10
		steps = [max_e]
		while True:
			next_value = steps[-1] - step_size
			if next_value < self.min_value() - self.delta_value(): break
			steps.append(next_value)
		for item in steps:
			try: value = int(round(450 - 400 * ((item - self.min_value()) / (self.delta_value()))))
			except ZeroDivisionError: value = 0;
			if 35 < value < 476:
				b = [
					'    <line x1="100" y1="{0}" x2="105" y2="{0}" stroke="black" stroke-width="2"/>',
					'    <text x="80" y="{}" text-anchor="middle" fill="black">{}</text>']
				item = self.commafy("{:.1f}".format(item))
				b[0] = b[0].format(value)
				b[1] = b[1].format(value,item)
				self.svg_code.extend(b)

	@functools.lru_cache(maxsize=1)
	def graph_crt_points(self):
		for i,opt in zip(self.paths,self.opt_atr):
			opt_cri = getattr(self.state, opt)[0]
			opt_con = getattr(self.state, opt)[1]
			if not len(i) == len(set([a[0] for a in i])):
				self.msg.append("WARNING: Two or more structures are occupying the same block lane!")
			l_c = [0, 0, 0]  # last collumn
			for idx, item in enumerate(i):
				c_p = [int((item[0]) * int(self.aesthetics["distance"]) + int(self.aesthetics["offset"]) + int(self.wide[0])), int(round(item[-1] + 50)),
					   int((item[0]) * int(self.aesthetics["distance"]) + int(self.aesthetics["offset"]) + self.wide[1])]
				# [x1,y1,x2], y1=y2
				a = [
					'    <line x1="{}" y1="{}" x2="{}" y2="{}" stroke="{}" stroke-width="{}" {}/>',
					'    <text x="{}" y="{}" text-anchor="middle" fill="{}">{}{}</text>',
					'    <text x="{}" y="{}" text-anchor="middle" fill="{}">{}</text>',
					'    <text x="{}" y="{}" text-anchor="middle" fill="{}">{}</text>']
				x = pref.svg_repl[opt_cri[-1]]
				z = pref.placement[item[-2]][0 if self.include_np else 1]
				trick_g = "g" if self.e_source == 3 else "h"
				trick_h = "h" if self.e_source == 3 else "g"
				digit_rounding = "{:.2f}".format(item[self.e_source]) if self.aesthetics["decimal"] == "2" else "{:.1f}".format(item[self.e_source])
				g = self.aesthetics[trick_g][0] + self.commafy(digit_rounding) + self.aesthetics[trick_g][-1]
				h = self.aesthetics[trick_h][0] + self.commafy(item[self.e_complement]) + self.aesthetics[trick_h][-1]
				ts_dict = {
				    " "        : "",
				    "‡ (big)":'<tspan dy="-7" font-family="arial" font-size=".7em">{}</tspan>'.format("‡".encode("ascii", "xmlcharrefreplace").decode("utf-8")),
				    "‡ (small)":'<tspan dy="-7" font-family="monospace" font-size=".7em">{}</tspan>'.format("‡".encode("ascii", "xmlcharrefreplace").decode("utf-8"))
				}
				ts_mark = ts_dict[self.aesthetics["mark"]] if item[1] == "TS" else ""
				a[0] = a[0].format(c_p[0], c_p[1], c_p[2], c_p[1], opt_cri[0],opt_cri[1],x)
				a[1] = a[1].format(int((c_p[0] + c_p[2])/2), c_p[1] + z[0], opt_cri[0],item[2].encode("ascii", "xmlcharrefreplace").decode("utf-8"),ts_mark)
				a[2] = a[2].format(int((c_p[0] + c_p[2])/2), c_p[1] + z[1],opt_cri[0],str(g).encode("ascii", "xmlcharrefreplace").decode("utf-8"))
				a[3] = a[3].format(int((c_p[0] + c_p[2])/2), c_p[1] + z[2],opt_cri[0],str(h).encode("ascii", "xmlcharrefreplace").decode("utf-8"))
				self.svg_code.extend(a if self.include_np else a[:-1])
				if not idx == 0:
					b = '    <line x1="{}" y1="{}" x2="{}" y2="{}" stroke="{}" stroke-width="{}" {}/>'
					x = pref.svg_repl[opt_con[-1]]
					b = b.format(l_c[2], l_c[1], c_p[0], c_p[1], opt_con[0],opt_con[1],x)
					self.svg_code.append(b)
				l_c = c_p
	@functools.lru_cache(maxsize=1)
	def graph_connectors(self):
		if not self.include[-1] == 1: return
		for i in getattr(self.state,"con"):
			if any(i[n] == "" for n in [0,1,2,3]):continue
			i = [int(a) if idx in [1, 3] else a for idx, a in enumerate(i)]
			data_dict = {a:b for a, b in zip(self.tab_v,pref.menu_e)}
			dict_a = {a:b for a,b in zip([data_dict[a] for a in self.data_atr],self.paths)}
			if not i[0] in dict_a.keys(): continue
			if not i[2] in dict_a.keys(): continue
			if not type(i[1]) == int: continue
			if not type(i[3]) == int: continue
			if not i[1] in [a[0] for a in dict_a[i[0]]]: continue
			if not i[3] in [a[0] for a in dict_a[i[2]]]: continue
			if i[1] == i[3]: self.msg.append("Cannot conect items on same column"); continue
			start = next(n for n in dict_a[i[0]] if n[0] == i[1])
			end = next(n for n in dict_a[i[2]] if n[0] == i[3])
			con = [start,end]
			if con[0][0] > con[1][0]:
				con.reverse()
			if con[0][0] < con[1][0]:
				x = pref.svg_repl[i[6]]
				a = '    <line x1="{}" y1="{}" x2="{}" y2="{}" stroke="{}" stroke-width="{}" {}/>'
				a = a.format(int((con[0][0]) * int(self.aesthetics["distance"]) + int(self.aesthetics["offset"]) + self.wide[1]), con[0][-1] + 50,
							 int((con[1][0]) * int(self.aesthetics["distance"]) + int(self.aesthetics["offset"]) + self.wide[0]), con[1][-1] + 50,
							 i[4],i[5],x)
				self.svg_code.append(a)

	@functools.lru_cache(maxsize=1)
	def span_dg(self):
		r_const = {"kcal/mol": 0.0019872, "kJ/mol": 0.0083144}[self.span["units"]]
		boltz_const = {"kcal/mol":3.29762e-27,"kJ/mol":1.380649e-26}[self.span["units"]]
		planck_const = {"kcal/mol":1.58367e-37,"kJ/mol":6.6260755e-37}[self.span["units"]]
		delta_e = self.paths[0][-1][self.e_source]-self.paths[0][0][self.e_source]
		limit = {"kcal/mol": 4, "kJ/mol": 16.736}[self.span["units"]]
		if not self.span_worthy: return
		if not len(self.paths[0]) > 1:
			self.msg.append("This software can only do span analysis if only one path is ploted\n")
			self.span_worthy = False; return
		if not len(self.set_height()) == 1: self.span_worthy = False; return
		z = "Analysis assumes structures #{} and #{} have the same geometry,"
		z += " but are energeticaly distiguished by the {} of the reaction.\n"
		self.msg.append(z.format(min(a[0] for a in self.paths[0]),
								 max(a[0] for a in self.paths[0]),
								 "exergonicity" if self.e_source == 3 else "exotermicity"))
		if self.e_source != 3:
			m = "WARNING: Data above should only be used after carefull consideration."
			m += "Enthalpy values were employed in place of Gibbs Free energy\n"
			self.msg.append(m)
		# Is it a TS or INT?
		if self.span["irrespective"] != 1:
			if not self.paths[0][0][1] == self.paths[0][-1][1]:
				self.msg.append("#{} and #{} must be the same TS/INT type\n".format(self.paths[0][0][0],self.paths[0][-1][0]))
			for i,a in enumerate(self.paths[0]):
				if i == 0 or i+1 == len(self.paths[0]):
					top = self.paths[0][1][self.e_source] < self.paths[0][0][self.e_source] and self.paths[0][-1][self.e_source] > self.paths[0][-2][self.e_source]
				else:
					top = self.paths[0][i-1][self.e_source] < a[self.e_source] > self.paths[0][i+1][self.e_source]
				if top and a[1] == "TS": pass
				elif top and a[1] == "INT": self.msg.append("Are you sure #{} is not a TS? It is directly connected to structures lower in both forward and backwards direction!\n".format(a[0]))
				elif not top and a[1] == "TS": self.msg.append("Are you sure #{} is not an INT? It is directly connected to structure(s) higher in energy!\n".format(a[0]))
				if a[1] not in ["TS","INT"]: self.msg.append("#{} should be set as either TS or INT, otherwise it will be ploted but excluded from analysis\n".format(a[0]))
		#TOF
		all_it = []
		for idx_a, a in enumerate(self.paths[0][:-1]):
			for idx_b,b in enumerate(self.paths[0][:-1]):
				all_it.append([a, a[self.e_source] - b[self.e_source] - delta_e if idx_b < idx_a else a[self.e_source] - b[self.e_source] , b])

		if self.span["irrespective"] != 1:
			all_it = [a for a in all_it if all([a[0][0] != a[2][0],a[0][1] == "TS", a[2][1] == "INT"])]
		if self.span["irrespective"] == 1:
			m = "WARNING: Data above should only be used after carefull consideration."
			m += "Equations were applied on the assumption that all structures are both intermediates and transition states simultaneously\n"
			self.msg.append(m)
		if all([self.span["irrespective"] != 1, all_it, self.e_source == 3]):
			self.msg.append("Ref.: Kozuch, S., Shaik, S. Acc. Chem. Res. 2011, 44, 101.\n")
		if all_it:
			denominator = sum(math.exp(a[1] / (self.temperature * r_const)) for a in all_it)
			tof = (((math.exp(-delta_e / (r_const * self.temperature)) - 1) / denominator) * self.temperature * boltz_const) / planck_const
			all_ts = list(dict.fromkeys([a[0][0] for a in all_it]))
			all_int = list(dict.fromkeys([a[2][0] for a in all_it]))
			x_tof_i = []
			x_tof_ts = []
			for x in all_int:
				a = sum(math.exp(a[1] / (self.temperature * r_const)) for a in all_it if x == a[2][0]) / denominator
				x_tof_i.append([x, a * 100])
			for x in all_ts:
				a = sum(math.exp(a[1] / (self.temperature * r_const)) for a in all_it if x == a[0][0]) / denominator
				x_tof_ts.append([x, a * 100])
			self.msg.append("".join("#{:>5}: {:>7.2f}% \n".format(*a) for a in x_tof_i))
			self.msg.append("X(tof) for intermediates:")
			self.msg.append("".join("#{:>5}: {:>7.2f}% \n".format(*a) for a in x_tof_ts))
			self.msg.append("X(tof) for transition states:")
			self.msg.append("TOF as catalytic flux law: {:5e} /h\n".format(tof * 3600))
			if abs(tof) > 1e8:
				self.msg.append("ALERT: Please consider the possibility of diffusion control rates\n")
		#CONDITIONALS FOR SPAN
		if delta_e >= 0:
			self.msg.append("Reaction is {}! Span will not be computed!\n".format("endergonic" if self.e_source == 3 else "endotermic"))
			self.span_worthy = False; return
		if not all(a[1] in ["TS","INT"] for a in self.paths[0]) and self.span["irrespective"] != 1:
			txt = "All structures have to be identified as either transition states or intermediates for a strict span analysis."
			txt += " Irrestrictive analysis may be caried out by checking the 'irrespective of type(TS/INT)' box."
			txt += " No span analysis will be conducted\n"
			self.msg.append(txt)
			self.span_worthy = False; return
		# SPAN
		all_it = []
		for idx_a, a in enumerate(self.paths[0][:-1]):
			for idx_b,b in enumerate(self.paths[0][:-1]):
				all_it.append([a, a[self.e_source] - b[self.e_source] + delta_e if idx_a <idx_b else a[self.e_source] - b[self.e_source], b])
		if self.span["irrespective"] != 1:
			all_it = [a for a in all_it if all([a[0][0] != a[2][0],a[0][1] == "TS",a[2][1] == "INT"])]
		#for a in all_it: print(a)
		if not all_it:
			self.span_worthy = False
			self.msg.append("No states found!\n")
			return
		span = max(all_it, key=lambda a: a[1])
		if span[1] <= 0:
			self.msg.append("The reaction appears to barrierless and therefore no span will be calculated!\n")
			self.span_worthy = False
			return
		for i,a in enumerate(all_it):
			if 0 <= a[1] > span[1] - limit and any(span[n][0] != a[n][0] for n in [0,2]):
				message = "WARNING: Span from #{} to #{} is only {:.2f} {} lower".format(a[0][0],a[2][0],span[1]-a[1],self.span["units"])
				message += " than #{} to #{} and may influence the rate determining state\n".format(span[0][0],span[2][0])
				self.msg.append(message)
		tof_span = (((math.exp(-span[1]/(r_const*self.temperature))))*self.temperature*boltz_const)/planck_const
		self.msg.append("TOF from span: {:5e} /h\n".format(tof_span*3600))
		return span

	@functools.lru_cache(maxsize=1)
	def graph_span(self):
		tdi_correct = {a: b for a,b in zip(pref.menu_d, [-35,-2,10] if self.include_np else [-20,-2,10])}
		tdts_correct = {a: b for a, b in zip(pref.menu_d, [-15, 18, 32] if self.include_np else [-15,2,16])}
		if self.span_dg() is None: self.span_worthy = False; return
		if self.span_worthy:
			data = [[self.span_dg()[0][n] for n in [0,-1]],[self.span_dg()[2][n] for n in [0,-1]]]
			span = self.span_dg()[1]
			data = sorted(data,key=lambda x: x[1])
			delta_e = self.paths[0][-1][self.e_source]-self.paths[0][0][self.e_source]
			if self.big_arrow:
				# TDI arrow big
				# print(self.data)
				x = tdi_correct[next(a for a in self.paths[0] if a[0] == data[1][0])[5]] - 40
				p = [(data[1][0]) * int(self.aesthetics["distance"]) + int(self.aesthetics["offset"]) + 10, data[1][1] + x]
				a = [
					'    <text x="{}" y="{}" text-anchor="middle" fill="black">TDI</text>',
					'    <path d=" M {0} {1} L {2} {1} L {2} {3} L {4} {3} L {5} {6} L {7} {3} L {0} {3} Z "/>']
				a[0] = a[0].format(p[0] + 20, p[1])
				a[1] = a[1].format(10 + p[0], 10 + p[1], 30 + p[0], 40 + p[1], 40 + p[0], 20 + p[0], 70 + p[1], 0 + p[0])
				self.svg_code.extend(a)
				# TDTS arrow
				x = tdts_correct[next(a for a in self.paths[0] if a[0] == data[0][0])[5]] + 140
				p = [(data[0][0]) * int(self.aesthetics["distance"]) + int(self.aesthetics["offset"]) + 50, data[0][1] + x]
				a = [
					'    <text x="{}" y="{}" text-anchor="middle" fill="black">TDTS</text>',
					'    <path d=" M {0} {1} L {2} {1} L {2} {3} L {4} {3} L {5} {6} L {7} {3} L {0} {3} Z "/>']
				a[0] = a[0].format(p[0] - 20, p[1] + 10)
				a[1] = a[1].format(-10 + p[0], -10 + p[1], -30 + p[0], -40 + p[1], -40 + p[0], -20 + p[0], -70 + p[1],
								   0 + p[0])
				self.svg_code.extend(a)
			else:
				# TDI arrow
				# print(self.data)
				x = tdi_correct[next(a for a in self.paths[0] if a[0] == data[1][0])[5]] - 40
				p = [(data[1][0]) * int(self.aesthetics["distance"]) + int(self.aesthetics["offset"]) + 10, data[1][1] + x]
				a = [
					'    <text x="{}" y="{}" text-anchor="middle" fill="black">TDI</text>',
					'    <text x="{}" y="{}" text-anchor="middle" fill="black">{}</text>']
				a[0] = a[0].format(p[0] + 20, p[1] + 45)
				a[1] = a[1].format(p[0] + 20, p[1] + 63,"↓".encode("ascii", "xmlcharrefreplace").decode("utf-8"))
				self.svg_code.extend(a)
				# TDTS arrow
				x = tdts_correct[next(a for a in self.paths[0] if a[0] == data[0][0])[5]] + 140
				p = [(data[0][0]) * int(self.aesthetics["distance"]) + int(self.aesthetics["offset"]) + 50, data[0][1] + x]
				a = [
					'    <text x="{}" y="{}" text-anchor="middle" fill="black">TDTS</text>',
					'    <text x="{}" y="{}" text-anchor="middle" fill="black">{}</text>']
				a[0] = a[0].format(p[0] - 20, p[1] -30)
				a[1] = a[1].format(p[0] - 20, p[1] -53,"↑".encode("ascii", "xmlcharrefreplace").decode("utf-8"))
				self.svg_code.extend(a)
			# dg and span anotations
			a = [
				'    <text x="120" y="450" text-anchor="left" fill="black">Delta = {}</text>',
				'    <text x="120" y="470" text-anchor="left" fill="black">Span = {}</text>']
			a[0] = a[0].format(self.commafy("{:.2f}".format(delta_e)))
			a[1] = a[1].format(self.commafy("{:.2f}".format(span)))
			self.svg_code.extend(a)

	@functools.lru_cache(maxsize=1)
	def return_svg_code(self):
		if not self.data_atr:
			self.msg.append("No data!");
			self.graph_frame();
			self.svg_code.append('</svg>');
			return self.svg_code
		self.set_height()
		self.graph_frame()
		self.graph_grid()
		self.graph_crt_points()
		self.graph_connectors()
		if self.span_request: self.span_dg()
		if self.span_request: self.graph_span()
		self.svg_code.append('</svg>')
		return self.svg_code
	def save_svg(self,svg_name):
		svg_name = ".E_profile.svg" if svg_name is None else svg_name
		try:
			with open(os.path.join(os.getcwd(),svg_name), "w") as out_file:
				for line in self.return_svg_code(): out_file.write(line + "\n")
			if svg_name != ".E_profile.svg":
				self.msg.append("Take a look at file {}!".format(svg_name))
			return self.msg
		except FileNotFoundError:
			pass

def initialize():
	global pref, note, window, frame2
	pref = Preferences()
	window = tk.Tk()
	#NOTEBOOK
	frame1 = tk.Frame(master=window)
	frame1.grid(column=0,row=0,rowspan=2,sticky="news")
	frame1.grid_columnconfigure(0, weight=1)
	frame1.grid_rowconfigure(0, weight=1)
	note = Note(frame1)
	#GENERAL
	frame2 = tk.Frame(master=window)
	frame2.grid(column=1,row=0,rowspan=1,sticky="news")
	frame2.grid_columnconfigure(0, weight=1)
	frame2.grid_rowconfigure(0, weight=1)
	frame3 = tk.Frame(master=window)
	label = tk.Label(frame3, text="github.com/ricalmang")
	label.grid(column=0,row=0,stick="news")
	frame3.grid(column=1,row=1,rowspan=1)
	menu = GeneralMenu(frame2,name="Actions")
	menu.grid(column=0, row=0,  columnspan=1, pady="0",padx="0", rowspan=1,sticky="news")
	window.grid_columnconfigure(0, weight=1)
	window.grid_rowconfigure(0, weight=1)
	w,h = 910, 685
	window.minsize(w,h)
	window.maxsize(2000,1200)
	ws = window.winfo_screenwidth() # width of the screen
	hs = window.winfo_screenheight() # height of the screen
	x = (ws/2) - (w/2)
	y = (hs/2) - (h/2)
	window.geometry('%dx%d+%d+%d' % (w, h, x, y))
	window.mainloop()
def is_str_float(i):
	try: float(i); return True
	except ValueError: return False
initialize()
